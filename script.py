import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from datetime import datetime, timedelta
from openpyxl import Workbook
from collections import defaultdict

# Função de autenticação na Google Search Console API
def authenticate_search_console():
    SCOPES = ['https://www.googleapis.com/auth/webmasters.readonly']
    SERVICE_ACCOUNT_FILE = '/Users/andersonmelo/Desktop/Scripts Phyton/valida_sitemap/automacao-seo-7afb76a54753.json'

    try:
        credentials = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('searchconsole', 'v1', credentials=credentials)
        return service
    except FileNotFoundError as e:
        print(f"Erro: Arquivo de credenciais não encontrado no caminho: {SERVICE_ACCOUNT_FILE}")
        raise e
    except Exception as e:
        print(f"Erro ao autenticar: {e}")
        raise e

# Função para obter as métricas de URLs em um período específico
def get_url_metrics(service, site_url, start_date, end_date):
    try:
        request = {
            'startDate': start_date,
            'endDate': end_date,
            'dimensions': ['page'],
            'rowLimit': 25000  # Limite de até 25.000 URLs
        }

        response = service.searchanalytics().query(siteUrl=site_url, body=request).execute()

        if 'rows' in response:
            # Retorna as URLs das páginas com cliques, impressões, CTR e posição
            return [{
                'url': row['keys'][0],
                'clicks': row['clicks'],
                'impressions': row['impressions'],
                'ctr': row['ctr'],
                'position': row['position']
            } for row in response['rows']]
        else:
            return []
    except Exception as e:
        print(f"Erro ao buscar métricas: {e}")
        return []

# Função para obter as métricas de palavras-chave em um período específico
def get_keyword_metrics(service, site_url, start_date, end_date):
    try:
        request = {
            'startDate': start_date,
            'endDate': end_date,
            'dimensions': ['query', 'page'],
            'rowLimit': 25000
        }

        response = service.searchanalytics().query(siteUrl=site_url, body=request).execute()

        if 'rows' in response:
            keyword_data = defaultdict(lambda: {'clicks': 0, 'impressions': 0, 'ctr': 0, 'position': 0, 'urls': []})
            for row in response['rows']:
                keyword = row['keys'][0]
                url = row['keys'][1]
                keyword_data[keyword]['clicks'] += row['clicks']
                keyword_data[keyword]['impressions'] += row['impressions']
                keyword_data[keyword]['ctr'] += row['ctr']
                keyword_data[keyword]['position'] += row['position']
                keyword_data[keyword]['urls'].append(url)

            # Ajusta CTR e posição para a média (dividido pelo número de URLs)
            for data in keyword_data.values():
                data['ctr'] /= len(data['urls'])
                data['position'] /= len(data['urls'])

            return keyword_data
        else:
            return {}
    except Exception as e:
        print(f"Erro ao buscar métricas de palavras-chave: {e}")
        return {}

# Função para unir as métricas de dois períodos (comparação entre atual e anterior)
def compare_metrics(current, previous):
    metrics_comparison = {}

    # Unir as métricas das duas listas com base na URL
    for url_info in current:
        url = url_info['url']
        metrics_comparison[url] = {
            'current_clicks': url_info['clicks'],
            'current_impressions': url_info['impressions'],
            'current_ctr': url_info['ctr'],
            'current_position': url_info['position'],
            'previous_clicks': 0,
            'previous_impressions': 0,
            'previous_ctr': 0,
            'previous_position': 0
        }

    for url_info in previous:
        url = url_info['url']
        if url in metrics_comparison:
            metrics_comparison[url]['previous_clicks'] = url_info['clicks']
            metrics_comparison[url]['previous_impressions'] = url_info['impressions']
            metrics_comparison[url]['previous_ctr'] = url_info['ctr']
            metrics_comparison[url]['previous_position'] = url_info['position']
        else:
            metrics_comparison[url] = {
                'current_clicks': 0,
                'current_impressions': 0,
                'current_ctr': 0,
                'current_position': 0,
                'previous_clicks': url_info['clicks'],
                'previous_impressions': url_info['impressions'],
                'previous_ctr': url_info['ctr'],
                'previous_position': url_info['position']
            }

    return metrics_comparison

# Função para calcular a variação percentual entre dois valores
def calculate_percentage_change(current, previous):
    if previous == 0:
        return "+100%" if current > 0 else "0%"  # 100% crescimento quando não havia valor anterior
    change = ((current - previous) / previous) * 100
    return f"+{change:.2f}%" if change > 0 else f"{change:.2f}%"

# Função para separar URLs por suas trilhas (pastas)
def separate_urls_by_trail(urls_indexed):
    trails = defaultdict(int)

    for url_info in urls_indexed:
        url = url_info
        trail = '/' + url.split('/', 3)[3].split('/')[0] + '/'
        trails[trail] += 1

    return trails

# Função para gerar um código sequencial único baseado nos arquivos existentes na pasta de saída
def generate_sequential_code(output_directory):
    existing_files = os.listdir(output_directory)
    report_files = [f for f in existing_files if f.startswith('urls_indexadas_relatorio_') and f.endswith('.xlsx')]

    # Encontrar o próximo código sequencial com base nos arquivos existentes
    if report_files:
        report_numbers = [int(f.split('_')[3]) for f in report_files if f.split('_')[3].isdigit()]
        next_number = max(report_numbers) + 1 if report_numbers else 1
    else:
        next_number = 1

    return next_number

# Função para realizar uma análise de quais URLs têm o melhor rendimento e onde investir
def analyze_best_performance(metrics_comparison):
    best_urls = sorted(metrics_comparison.items(), key=lambda item: item[1]['current_clicks'], reverse=True)[:10]
    return best_urls

# Função para listar todas as URLs indexadas com métricas, salvar em planilha e exibir relatório analítico
def list_indexed_urls(service, site_url, output_file):
    # Definir datas dos últimos 30 dias e 30 dias anteriores
    end_date = datetime.now()
    start_date_current = (end_date - timedelta(days=30)).strftime('%Y-%m-%d')
    end_date_current = end_date.strftime('%Y-%m-%d')
    start_date_previous = (end_date - timedelta(days=60)).strftime('%Y-%m-%d')
    end_date_previous = (end_date - timedelta(days=31)).strftime('%Y-%m-%d')

    # Obter as métricas dos últimos 30 dias
    metrics_current = get_url_metrics(service, site_url, start_date_current, end_date_current)
    # Obter as métricas dos 30 dias anteriores
    metrics_previous = get_url_metrics(service, site_url, start_date_previous, end_date_previous)

    # Comparar os dois períodos
    metrics_comparison = compare_metrics(metrics_current, metrics_previous)

    if not metrics_comparison:
        print("Nenhuma URL indexada encontrada.")
        return

    total_urls = len(metrics_comparison)

    # Separar URLs por trilhas/pastas
    trails = separate_urls_by_trail(metrics_comparison.keys())

    # Relatório analítico na tela (MU Editor)
    print("\n--- Relatório Analítico de Indexação (Comparação Últimos 30 Dias) ---")
    print(f"Data de Execução: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Total de URLs Indexadas Encontradas: {total_urls}")
    print(f"Taxa de Indexação: {total_urls} URLs indexadas")
    print("\n--- URLs Indexadas por Trilha ---")
    for trail, count in trails.items():
        print(f"{trail}: {count} páginas indexadas")
    print("-----------------------------------------\n")

    # Criar uma planilha Excel com as URLs e informações analíticas
    wb = Workbook()

    # Adicionar uma aba de resumo com as informações analíticas
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_summary.append(["Resumo da Indexação"])
    ws_summary.append(["Data de Execução", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_summary.append(["Total de URLs Indexadas", total_urls])
    ws_summary.append(["Taxa de Indexação"])

    for trail, count in trails.items():
        ws_summary.append([trail, f"{count} páginas indexadas"])

    # Adicionar uma aba com as URLs indexadas e suas métricas (lado a lado e com evolução)
    ws_urls = wb.create_sheet("URLs Indexadas")
    ws_urls.append([
        "URL",
        "Cliques (Últimos 30 Dias)", "Cliques (30 Dias Anteriores)", "Evolução de Cliques",
        "Impressões (Últimos 30 Dias)", "Impressões (30 Dias Anteriores)", "Evolução de Impressões",
        "CTR (Últimos 30 Dias)", "CTR (30 Dias Anteriores)", "Evolução de CTR",
        "Posição Média (Últimos 30 Dias)", "Posição Média (30 Dias Anteriores)", "Evolução de Posição"
    ])

    for url, data in metrics_comparison.items():
        ws_urls.append([
            url,
            data['current_clicks'], data['previous_clicks'], calculate_percentage_change(data['current_clicks'], data['previous_clicks']),
            data['current_impressions'], data['previous_impressions'], calculate_percentage_change(data['current_impressions'], data['previous_impressions']),
            f"{data['current_ctr'] * 100:.2f}%", f"{data['previous_ctr'] * 100:.2f}%", calculate_percentage_change(data['current_ctr'], data['previous_ctr']),
            f"{data['current_position']:.2f}", f"{data['previous_position']:.2f}", calculate_percentage_change(data['previous_position'], data['current_position'])  # Invertendo a posição
        ])

    # Adicionar uma aba com as trilhas e a contagem de páginas indexadas
    ws_trails = wb.create_sheet("URLs por Trilha")
    ws_trails.append(["Trilha", "Quantidade de Páginas Indexadas"])
    for trail, count in trails.items():
        ws_trails.append([trail, count])

    # Adicionar uma aba com as URLs com melhor rendimento (análise de desempenho)
    ws_performance = wb.create_sheet("Melhores URLs")
    ws_performance.append(["URL", "Cliques (Últimos 30 Dias)", "Impressões (Últimos 30 Dias)", "CTR", "Posição Média"])
    best_urls = analyze_best_performance(metrics_comparison)
    for url, data in best_urls:
        ws_performance.append([
            url,
            data['current_clicks'],
            data['current_impressions'],
            f"{data['current_ctr'] * 100:.2f}%",
            f"{data['current_position']:.2f}"
        ])

    # Adicionar uma aba de palavras-chave e URLs
    start_date_current = (end_date - timedelta(days=30)).strftime('%Y-%m-%d')
    keyword_metrics = get_keyword_metrics(service, site_url, start_date_current, end_date_current)

    ws_keywords = wb.create_sheet("Palavras-Chave Indexadas")
    ws_keywords.append(["Palavra-Chave", "Cliques", "Impressões", "CTR", "Posição Média", "URLs Associadas"])

    for keyword, data in keyword_metrics.items():
        row = [keyword, data['clicks'], data['impressions'], f"{data['ctr'] * 100:.2f}%", f"{data['position']:.2f}"]
        row.extend(data['urls'])  # Adiciona as URLs associadas à palavra-chave
        ws_keywords.append(row)

    # Adicionar uma aba com as melhores palavras-chave e URLs para trabalhar
    ws_best_keywords = wb.create_sheet("Melhores Palavras-Chave")
    best_keywords = sorted(keyword_metrics.items(), key=lambda x: x[1]['clicks'], reverse=True)[:10]
    ws_best_keywords.append(["Palavra-Chave", "Cliques", "Impressões", "CTR", "Posição Média", "URLs Associadas"])

    for keyword, data in best_keywords:
        row = [keyword, data['clicks'], data['impressions'], f"{data['ctr'] * 100:.2f}%", f"{data['position']:.2f}"]
        row.extend(data['urls'])  # Adiciona as URLs associadas à palavra-chave
        ws_best_keywords.append(row)

    # Salvar a planilha
    try:
        wb.save(output_file)
        print(f"\nRelatório de URLs e palavras-chave salvo em: {output_file}")
    except Exception as e:
        print(f"Erro ao salvar as URLs e palavras-chave no arquivo: {e}")

# Função principal
def main():
    # Solicitar a URL do site (prefixo completo)
    site_url = 'https://advogadospelobrasil.com.br'  # Defina seu domínio (prefixo completo)

    # Autenticação
    service = authenticate_search_console()

    # Diretório de saída
    output_directory = os.getcwd()

    # Gerar o código sequencial único
    code = generate_sequential_code(output_directory)

    # Nome do arquivo de saída (planilha) com data, hora e código sequencial
    output_file = os.path.join(output_directory, f'urls_indexadas_relatorio_{code}_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')

    # Listar todas as URLs indexadas e salvar na planilha
    list_indexed_urls(service, site_url, output_file)

if __name__ == '__main__':
    main()
